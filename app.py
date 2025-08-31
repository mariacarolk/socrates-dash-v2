#!/usr/bin/env python3
"""
Sócrates Online - Aplicação Web Flask
Importação de Relatório de Faturamento de Eventos
"""

import os
import json
import io
import base64
import csv
import threading
import time
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
import pandas as pd
# import plotly.express as px
# import plotly.graph_objects as go
# import plotly.utils
try:
    import plotly.express as px
    import plotly.graph_objects as go
    import plotly.utils
    PLOTLY_AVAILABLE = True
    print("✅ Plotly carregado com sucesso")
except ImportError as e:
    print(f"⚠️ Plotly não disponível: {e}")
    PLOTLY_AVAILABLE = False
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import re

app = Flask(__name__)
app.secret_key = 'socrates_sistema_2025_secret_key'

# Configurações
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB máx

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Criar pasta de uploads se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

class CircosCidadesManager:
    """Classe para gerenciar dados de circos e cidades"""
    
    def __init__(self, csv_file='circos_cidades.csv'):
        self.csv_file = csv_file
        self.backup_folder = 'backups_csv'
        self.data = []
        self._ensure_backup_folder()
        self.load_data()
    
    def _ensure_backup_folder(self):
        """Criar pasta de backup e mover arquivos existentes"""
        try:
            # Criar pasta de backup se não existir
            os.makedirs(self.backup_folder, exist_ok=True)
            print(f"📁 Pasta de backup verificada: {os.path.abspath(self.backup_folder)}")
            
            # Mover backups existentes da raiz para a nova pasta
            self._move_existing_backups()
            
        except Exception as e:
            print(f"⚠️ Erro ao configurar pasta de backup: {e}")
    
    def _move_existing_backups(self):
        """Mover backups existentes da raiz para a pasta de backup"""
        try:
            moved_count = 0
            
            # Procurar arquivos de backup na raiz
            for filename in os.listdir('.'):
                if (filename.startswith(f"{self.csv_file}.backup") and 
                    filename != self.csv_file):
                    
                    source_path = filename
                    dest_path = os.path.join(self.backup_folder, filename)
                    
                    # Verificar se o arquivo já existe no destino
                    if not os.path.exists(dest_path):
                        import shutil
                        shutil.move(source_path, dest_path)
                        moved_count += 1
                        print(f"📦 Backup movido: {filename} → {self.backup_folder}/")
                    else:
                        # Se já existe, remover o da raiz
                        os.remove(source_path)
                        print(f"🗑️ Backup duplicado removido da raiz: {filename}")
            
            if moved_count > 0:
                print(f"✅ {moved_count} arquivos de backup organizados na pasta {self.backup_folder}")
            
        except Exception as e:
            print(f"⚠️ Erro ao mover backups existentes: {e}")
    
    def load_data(self):
        """Carregar dados do CSV"""
        try:
            if os.path.exists(self.csv_file):
                print(f"📁 Carregando dados de: {os.path.abspath(self.csv_file)}")
                with open(self.csv_file, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    self.data = list(reader)
                print(f"✅ Dados de circos carregados: {len(self.data)} registros")
                
                # Debug: mostrar primeiros registros
                if self.data:
                    print(f"📋 Primeira entrada: {self.data[0]}")
            else:
                print(f"⚠️ Arquivo CSV não encontrado: {os.path.abspath(self.csv_file)}")
                print("📝 Criando arquivo vazio...")
                self.data = []
                self.save_data()  # Criar arquivo vazio
        except Exception as e:
            print(f"❌ ERRO ao carregar dados de circos: {e}")
            import traceback
            traceback.print_exc()
            self.data = []
    
    def backup_data(self):
        """Criar backup do arquivo antes de salvar"""
        try:
            if os.path.exists(self.csv_file):
                import shutil
                
                # Backup principal na pasta de backups
                backup_filename = f"{os.path.basename(self.csv_file)}.backup"
                backup_file = os.path.join(self.backup_folder, backup_filename)
                shutil.copy2(self.csv_file, backup_file)
                
                # Backup com timestamp para histórico
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                timestamped_filename = f"{os.path.basename(self.csv_file)}.backup_{timestamp}"
                timestamped_backup = os.path.join(self.backup_folder, timestamped_filename)
                shutil.copy2(self.csv_file, timestamped_backup)
                
                print(f"🔄 Backup criado: {backup_file}")
                print(f"🔄 Backup timestamped: {timestamped_backup}")
                
                # Limpar backups antigos (manter apenas os últimos 10)
                self.cleanup_old_backups()
                
        except Exception as e:
            print(f"⚠️ Falha ao criar backup: {e}")
    
    def cleanup_old_backups(self):
        """Limpar backups antigos, mantendo apenas os últimos 10"""
        try:
            backup_pattern = f"{os.path.basename(self.csv_file)}.backup_"
            backup_files = []
            
            # Encontrar todos os backups timestamped na pasta de backup
            for filename in os.listdir(self.backup_folder):
                if filename.startswith(backup_pattern):
                    full_path = os.path.join(self.backup_folder, filename)
                    backup_files.append((full_path, os.path.getmtime(full_path)))
            
            # Ordenar por data de modificação (mais recente primeiro)
            backup_files.sort(key=lambda x: x[1], reverse=True)
            
            # Manter apenas os 10 mais recentes
            for backup_file, _ in backup_files[10:]:
                try:
                    os.remove(backup_file)
                    print(f"🗑️ Backup antigo removido: {backup_file}")
                except Exception as e:
                    print(f"⚠️ Erro ao remover backup antigo {backup_file}: {e}")
                    
        except Exception as e:
            print(f"⚠️ Erro na limpeza de backups: {e}")
    
    def auto_save_thread(self):
        """Thread para salvamento automático periódico"""
        while getattr(self, '_auto_save_active', True):
            try:
                time.sleep(300)  # Salvar a cada 5 minutos
                if self.data:  # Só salvar se há dados
                    print(f"💾 Salvamento automático...")
                    self.save_data()
            except Exception as e:
                print(f"⚠️ Erro no salvamento automático: {e}")
    
    def start_auto_save(self):
        """Iniciar thread de salvamento automático"""
        try:
            self._auto_save_active = True
            self._auto_save_thread = threading.Thread(target=self.auto_save_thread, daemon=True)
            self._auto_save_thread.start()
            print("⏰ Salvamento automático iniciado (a cada 5 minutos)")
        except Exception as e:
            print(f"⚠️ Erro ao iniciar salvamento automático: {e}")
    
    def stop_auto_save(self):
        """Parar thread de salvamento automático"""
        try:
            self._auto_save_active = False
            print("⏹️ Salvamento automático interrompido")
        except Exception as e:
            print(f"⚠️ Erro ao parar salvamento automático: {e}")

    def save_data(self):
        """Salvar dados no CSV"""
        try:
            # Criar backup antes de salvar
            self.backup_data()
            
            # Sempre criar o arquivo, mesmo que vazio
            with open(self.csv_file, 'w', newline='', encoding='utf-8') as file:
                fieldnames = ['CIDADE', 'CIRCO', 'DATA_INICIO', 'DATA_FIM']
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                
                # Salvar dados se existirem
                if self.data:
                    writer.writerows(self.data)
                    
            print(f"✅ Dados de circos salvos: {len(self.data)} registros")
            print(f"📁 Arquivo: {os.path.abspath(self.csv_file)}")
            
            # Verificar se o arquivo foi realmente criado
            if os.path.exists(self.csv_file):
                file_size = os.path.getsize(self.csv_file)
                print(f"📊 Tamanho do arquivo: {file_size} bytes")
                
                # Verificar integridade dos dados salvos
                try:
                    with open(self.csv_file, 'r', encoding='utf-8') as verify_file:
                        verify_reader = csv.DictReader(verify_file)
                        verify_data = list(verify_reader)
                        if len(verify_data) == len(self.data):
                            print(f"✅ Verificação de integridade: OK ({len(verify_data)} registros)")
                        else:
                            print(f"⚠️ Integridade comprometida: esperado {len(self.data)}, encontrado {len(verify_data)}")
                except Exception as verify_error:
                    print(f"⚠️ Erro na verificação: {verify_error}")
            else:
                print("⚠️ ALERTA: Arquivo não foi criado!")
                return False
                
            return True
        except Exception as e:
            print(f"❌ ERRO CRÍTICO ao salvar dados de circos: {e}")
            print(f"📁 Tentando salvar em: {os.path.abspath(self.csv_file)}")
            import traceback
            traceback.print_exc()
            return False
    
    def add_circo(self, cidade, circo, data_inicio, data_fim):
        """Adicionar novo registro"""
        try:
            novo_registro = {
                'CIDADE': cidade,
                'CIRCO': circo,
                'DATA_INICIO': data_inicio,
                'DATA_FIM': data_fim
            }
            
            # Backup dos dados atuais antes da modificação
            dados_backup = self.data.copy()
            
            self.data.append(novo_registro)
            success = self.save_data()
            
            if success:
                # Validar que os dados foram salvos corretamente
                if self.validate_data_integrity():
                    print(f"✅ Circo adicionado e validado: {circo} em {cidade}")
                    return True
                else:
                    # Reverter se a validação falhou
                    self.data = dados_backup
                    self.save_data()
                    print(f"❌ Falha na validação após adicionar circo, dados revertidos")
                    return False
            else:
                # Reverter se o salvamento falhou
                self.data = dados_backup
                print(f"❌ Falha no salvamento, dados revertidos")
                return False
                
        except Exception as e:
            print(f"❌ Erro ao adicionar circo: {e}")
            return False
    
    def update_circo(self, index, cidade, circo, data_inicio, data_fim):
        """Atualizar registro existente"""
        if 0 <= index < len(self.data):
            try:
                # Backup dos dados atuais
                dados_backup = self.data.copy()
                
                self.data[index] = {
                    'CIDADE': cidade,
                    'CIRCO': circo,
                    'DATA_INICIO': data_inicio,
                    'DATA_FIM': data_fim
                }
                
                success = self.save_data()
                
                if success and self.validate_data_integrity():
                    print(f"✅ Circo atualizado e validado: {circo} em {cidade}")
                    return True
                else:
                    # Reverter se falhou
                    self.data = dados_backup
                    self.save_data()
                    print(f"❌ Falha na atualização, dados revertidos")
                    return False
                    
            except Exception as e:
                print(f"❌ Erro ao atualizar circo: {e}")
                return False
        return False
    
    def delete_circo(self, index):
        """Deletar registro"""
        if 0 <= index < len(self.data):
            try:
                # Backup dos dados atuais
                dados_backup = self.data.copy()
                deleted_item = self.data[index].copy()
                
                del self.data[index]
                success = self.save_data()
                
                if success and self.validate_data_integrity():
                    print(f"✅ Circo deletado e validado: {deleted_item['CIRCO']} em {deleted_item['CIDADE']}")
                    return True
                else:
                    # Reverter se falhou
                    self.data = dados_backup
                    self.save_data()
                    print(f"❌ Falha na deleção, dados revertidos")
                    return False
                    
            except Exception as e:
                print(f"❌ Erro ao deletar circo: {e}")
                return False
        return False
    
    def get_all(self):
        """Obter todos os registros"""
        return self.data
    
    def get_circos_unicos(self):
        """Obter lista de circos únicos"""
        circos = set(item['CIRCO'] for item in self.data)
        return sorted(list(circos))
    
    def verify_and_recover(self):
        """Verificar dados e tentar recuperar do backup se necessário"""
        try:
            print(f"🔍 Verificando integridade dos dados...")
            
            # Verificar se o arquivo principal existe e é válido
            if not os.path.exists(self.csv_file):
                print(f"⚠️ Arquivo principal não encontrado: {self.csv_file}")
                return self.recover_from_backup()
            
            # Verificar se o arquivo tem conteúdo válido
            try:
                with open(self.csv_file, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    test_data = list(reader)
                    if not test_data and len(self.data) > 0:
                        print("⚠️ Arquivo principal vazio, mas dados em memória existem")
                        return self.recover_from_backup()
            except Exception as e:
                print(f"⚠️ Erro ao ler arquivo principal: {e}")
                return self.recover_from_backup()
            
            print("✅ Dados íntegros")
            return True
            
        except Exception as e:
            print(f"❌ Erro na verificação: {e}")
            return False
    
    def recover_from_backup(self):
        """Recuperar dados do backup"""
        try:
            backup_filename = f"{os.path.basename(self.csv_file)}.backup"
            backup_file = os.path.join(self.backup_folder, backup_filename)
            if os.path.exists(backup_file):
                print(f"🔄 Recuperando dados do backup: {backup_file}")
                import shutil
                shutil.copy2(backup_file, self.csv_file)
                self.load_data()
                print(f"✅ Dados recuperados: {len(self.data)} registros")
                return True
            else:
                # Tentar recuperar do backup mais recente com timestamp
                return self.recover_from_timestamped_backup()
        except Exception as e:
            print(f"❌ Erro ao recuperar backup: {e}")
            return False
    
    def recover_from_timestamped_backup(self):
        """Recuperar do backup timestamped mais recente"""
        try:
            backup_pattern = f"{os.path.basename(self.csv_file)}.backup_"
            backup_files = []
            
            for filename in os.listdir(self.backup_folder):
                if filename.startswith(backup_pattern):
                    full_path = os.path.join(self.backup_folder, filename)
                    backup_files.append((full_path, os.path.getmtime(full_path)))
            
            if backup_files:
                # Pegar o backup mais recente
                backup_files.sort(key=lambda x: x[1], reverse=True)
                most_recent_backup = backup_files[0][0]
                
                print(f"🔄 Recuperando do backup timestamped: {most_recent_backup}")
                import shutil
                shutil.copy2(most_recent_backup, self.csv_file)
                self.load_data()
                print(f"✅ Dados recuperados do backup timestamped: {len(self.data)} registros")
                return True
            else:
                print("❌ Nenhum backup encontrado")
                return False
                
        except Exception as e:
            print(f"❌ Erro ao recuperar backup timestamped: {e}")
            return False
    
    def validate_data_integrity(self):
        """Validar integridade dos dados salvos"""
        try:
            if not os.path.exists(self.csv_file):
                print("❌ Arquivo não existe após salvamento")
                return False
            
            # Ler arquivo e comparar com dados em memória
            with open(self.csv_file, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                file_data = list(reader)
            
            if len(file_data) != len(self.data):
                print(f"❌ Divergência no número de registros: arquivo={len(file_data)}, memória={len(self.data)}")
                return False
            
            # Verificar se os dados são equivalentes (ordem pode variar)
            for mem_item in self.data:
                found = False
                for file_item in file_data:
                    if (mem_item['CIDADE'] == file_item['CIDADE'] and 
                        mem_item['CIRCO'] == file_item['CIRCO'] and
                        mem_item['DATA_INICIO'] == file_item['DATA_INICIO'] and
                        mem_item['DATA_FIM'] == file_item['DATA_FIM']):
                        found = True
                        break
                
                if not found:
                    print(f"❌ Registro não encontrado no arquivo: {mem_item}")
                    return False
            
            print("✅ Integridade dos dados validada")
            return True
            
        except Exception as e:
            print(f"❌ Erro na validação de integridade: {e}")
            return False

class SocratesProcessor:
    """Classe para processar dados do Sócrates Online"""
    
    def __init__(self):
        self.processed_data = []
        self.original_df = None
    
    def allowed_file(self, filename):
        """Verifica se o arquivo é permitido"""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
    def extract_circo_name(self, evento_text):
        """Extrai o nome do circo do texto do evento"""
        
        # Verificar se é NAN ou inválido
        if pd.isna(evento_text) or str(evento_text).lower() in ['nan', 'none', '']:
            return "Evento Inválido"
        
        evento = str(evento_text).strip()
        
        # Caso 1: Se há barra |, usar texto antes da barra
        if '|' in evento:
            return evento.split('|')[0].strip()
        
        # Caso 2: Se não há barra, procurar por padrões de data e usar texto antes da data
        # Incluindo dias da semana para remoção
        
        # Dias da semana em português
        dias_semana = r'(?:segunda|terça|terca|quarta|quinta|sexta|sábado|sabado|domingo|seg|ter|qua|qui|sex|sab|dom)'
        
        # Padrões mais específicos incluindo dias da semana
        date_patterns = [
            # Padrão: "Domingo 16.MAR" ou "domingo 16.mar"
            rf'\s+{dias_semana}\s+\d{{1,2}}\.\w{{3}}',
            # Padrão: "Domingo 16/03" 
            rf'\s+{dias_semana}\s+\d{{1,2}}/\d{{1,2}}',
            # Padrão: "16.MAR"
            r'\s+\d{1,2}\.\w{3}',
            # Padrão: "16/03"
            r'\s+\d{1,2}/\d{1,2}',
            # Padrão: "16-03"
            r'\s+\d{1,2}-\d{1,2}',
            # Padrão: "16 MAR"
            r'\s+\d{1,2}\s+\w{3}',
            # Padrão apenas dia da semana no final
            rf'\s+{dias_semana}$',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, evento, re.IGNORECASE)
            if match:
                # Retornar texto antes do padrão encontrado
                circo_name = evento[:match.start()].strip()
                if circo_name and len(circo_name) > 2:  # Nome deve ter pelo menos 3 caracteres
                    return circo_name
        
        # Caso 3: Se não encontrou padrão, remover dias da semana do final
        # Padrão para remover dias da semana soltos no final
        evento_limpo = re.sub(rf'\s+{dias_semana}$', '', evento, flags=re.IGNORECASE)
        if evento_limpo != evento and len(evento_limpo.strip()) > 2:
            return evento_limpo.strip()
        
        # Caso 4: Se não encontrou padrão de data, retornar o texto completo (se válido)
        if len(evento.strip()) > 2:
            return evento.strip()
        
        return "Evento Sem Nome"

    def format_currency(self, value):
        """Formata valores monetários"""
        try:
            if pd.isna(value):
                return 0
            
            # Converter para float se for string
            if isinstance(value, str):
                # Remover caracteres não numéricos exceto vírgula e ponto
                value = value.replace('R$', '').replace('.', '').replace(',', '.')
                value = ''.join(c for c in value if c.isdigit() or c == '.')
                value = float(value) if value else 0
            
            return float(value)
            
        except:
            return 0

    def format_currency_display(self, value):
        """Formata valores para exibição"""
        try:
            return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "R$ 0,00"

    def process_excel_file(self, file_path):
        """Processa arquivo Excel e retorna dados processados"""
        try:
            # Ler arquivo Excel
            df = pd.read_excel(file_path)
            self.original_df = df.copy()
            
            # Verificar colunas necessárias
            required_columns = ['Evento', 'Data Evento', 'Faturamento Total', 'Faturamento Gestão Produtor']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return False, f"Colunas não encontradas: {', '.join(missing_columns)}"
            
            # Processar dados
            self.processed_data = []
            
            for index, row in df.iterrows():
                try:
                    # Verificar se o evento é válido (não NAN)
                    if pd.isna(row['Evento']):
                        continue
                    
                    # Processar coluna Evento - extrair nome do circo
                    evento = str(row['Evento'])
                    circo = self.extract_circo_name(evento)
                    
                    # Pular registros com nomes inválidos
                    if circo in ['Evento Inválido', 'Evento Sem Nome']:
                        continue
                    
                    # Processar Data Evento
                    data_evento = row['Data Evento']
                    if pd.isna(data_evento):
                        data_evento = "Não informado"
                    else:
                        # Tentar formatar a data com formato brasileiro
                        try:
                            if isinstance(data_evento, str):
                                # Tentar diferentes formatos de data brasileira
                                for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%d/%m/%y']:
                                    try:
                                        parsed_date = datetime.strptime(data_evento, fmt)
                                        data_evento = parsed_date.strftime('%d/%m/%Y')
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    # Se nenhum formato brasileiro funcionou, tentar pd.to_datetime com formato brasileiro forçado
                                    try:
                                        # Forçar formato brasileiro DD/MM/YYYY
                                        parsed_date = pd.to_datetime(data_evento, format='%d/%m/%Y')
                                        data_evento = parsed_date.strftime('%d/%m/%Y')
                                    except:
                                        try:
                                            # Tentar interpretação com dayfirst=True (formato brasileiro)
                                            parsed_date = pd.to_datetime(data_evento, dayfirst=True)
                                            data_evento = parsed_date.strftime('%d/%m/%Y')
                                        except:
                                            data_evento = str(data_evento)
                            else:
                                # Se for datetime, apenas reformatar
                                data_evento = data_evento.strftime('%d/%m/%Y')
                        except Exception as e:
                            print(f"⚠️ Erro ao processar data '{data_evento}': {e}")
                            data_evento = str(data_evento)
                    
                    # Processar valores de faturamento
                    faturamento_total = self.format_currency(row['Faturamento Total'])
                    faturamento_gestao = self.format_currency(row['Faturamento Gestão Produtor'])
                    
                    # Processar taxas e descontos
                    taxas_columns = [
                        'Taxa Antecipação',
                        'Taxa Transferencia',
                        'I:Comissão Bilheteria e PDVS',
                        'I:Insumo - Ingresso Cancelado',
                        'I:Insumo - Ingresso Cortesia',
                        'I:Taxas Cartões - Debito',
                        'I:Taxas Cartões - Credito à Vista',
                        'I:Taxa Pix',
                        'I:Despesas Jurídicas'
                    ]
                    
                    taxas_e_descontos = 0
                    for col in taxas_columns:
                        if col in row and not pd.isna(row[col]):
                            taxas_e_descontos += self.format_currency(row[col])
                    
                    # Calcular novo valor líquido (Total - Gestão Produtor - Taxas e Descontos)
                    valor_liquido = faturamento_total - faturamento_gestao - taxas_e_descontos
                    
                    # Adicionar aos dados processados
                    self.processed_data.append({
                        'Circo': circo,
                        'Data Evento': data_evento,
                        'Evento Completo': evento,
                        'Faturamento Total': faturamento_total,
                        'Faturamento Gestão Produtor': faturamento_gestao,
                        'Taxas e Descontos': taxas_e_descontos,
                        'Valor Líquido': valor_liquido
                    })
                    
                except Exception as e:
                    continue
            
            return True, f"{len(self.processed_data)} registros processados com sucesso"
            
        except Exception as e:
            return False, f"Erro ao processar arquivo: {str(e)}"

    def filter_and_generate_report(self, selected_circos, data_inicio, data_fim):
        """Filtra dados e gera relatório"""
        
        # Converter dados processados para DataFrame
        df = pd.DataFrame(self.processed_data)
        
        if df.empty:
            return []
        
        # Filtrar por circos selecionados
        df = df[df['Circo'].isin(selected_circos)]
        
        # Converter e filtrar por data (formato brasileiro)
        try:
            # Forçar interpretação brasileira das datas
            df['Data Evento'] = pd.to_datetime(df['Data Evento'], format='%d/%m/%Y', errors='coerce')
            df = df.dropna(subset=['Data Evento'])
            
            # Filtrar por período
            df = df[
                (df['Data Evento'].dt.date >= data_inicio) & 
                (df['Data Evento'].dt.date <= data_fim)
            ]
        except Exception as e:
            print(f"⚠️ Erro no filtro de datas: {e}")
            pass
        
        if df.empty:
            return []
        
        # Agrupar por circo e somar faturamentos
        grouped = df.groupby('Circo').agg({
            'Faturamento Total': 'sum',
            'Faturamento Gestão Produtor': 'sum',
            'Taxas e Descontos': 'sum',
            'Valor Líquido': 'sum'
        }).reset_index()
        
        # Criar dados do relatório
        report_data = []
        periodo_str = f"{data_inicio.strftime('%d/%m/%Y')} - {data_fim.strftime('%d/%m/%Y')}"
        
        for _, row in grouped.iterrows():
            report_data.append({
                'Circo': row['Circo'],
                'Período': periodo_str,
                'Faturamento Total': row['Faturamento Total'],
                'Faturamento Gestão Produtor': row['Faturamento Gestão Produtor'],
                'Taxas e Descontos': row['Taxas e Descontos'],
                'Valor Líquido': row['Valor Líquido'],
                'Total Geral': row['Faturamento Total']
            })
        
        # Armazenar dados do relatório para exportação
        self.last_report_data = report_data
        print(f"Relatório gerado com {len(report_data)} registros")
        return report_data
    
    def filter_and_generate_report_with_cities(self, selected_circos, selected_cidades, data_inicio, data_fim):
        """Filtrar dados e gerar relatório incluindo filtro por cidades"""
        if not self.processed_data:
            return []
        
        # Converter para DataFrame
        df = pd.DataFrame(self.processed_data)
        
        # Filtrar por circos
        df = df[df['Circo'].isin(selected_circos)]
        
        # Filtrar por período (formato brasileiro)
        try:
            df['Data Evento Parsed'] = pd.to_datetime(df['Data Evento'], format='%d/%m/%Y', errors='coerce')
            df = df.dropna(subset=['Data Evento Parsed'])
            df = df[
                (df['Data Evento Parsed'].dt.date >= data_inicio) &
                (df['Data Evento Parsed'].dt.date <= data_fim)
            ]
        except Exception as e:
            print(f"⚠️ Erro no filtro de datas (com cidades): {e}")
            pass
        
        if df.empty:
            return []
        
        # Fazer associação com cidades para filtrar
        circos_cidades = circos_manager.get_all()
        
        # Adicionar coluna de cidade aos dados
        df['Cidade'] = 'Não encontrada'
        
        for index, row in df.iterrows():
            circo = row['Circo']
            data_evento_str = row['Data Evento']
            
            try:
                data_evento = datetime.strptime(data_evento_str, '%d/%m/%Y').date()
            except:
                continue
            
            for circo_cidade in circos_cidades:
                if circo_cidade['CIRCO'] == circo:
                    try:
                        data_inicio_cc = datetime.strptime(circo_cidade['DATA_INICIO'], '%d/%m/%Y').date()
                        data_fim_cc = datetime.strptime(circo_cidade['DATA_FIM'], '%d/%m/%Y').date()
                        
                        if data_inicio_cc <= data_evento <= data_fim_cc:
                            df.at[index, 'Cidade'] = circo_cidade['CIDADE']
                            break
                    except:
                        continue
        
        # Filtrar por cidades selecionadas
        df = df[df['Cidade'].isin(selected_cidades)]
        
        if df.empty:
            return []
        
        # Agrupar por circo e somar faturamentos
        grouped = df.groupby('Circo').agg({
            'Faturamento Total': 'sum',
            'Faturamento Gestão Produtor': 'sum',
            'Taxas e Descontos': 'sum',
            'Valor Líquido': 'sum'
        }).reset_index()
        
        # Criar dados do relatório
        report_data = []
        periodo_str = f"{data_inicio.strftime('%d/%m/%Y')} - {data_fim.strftime('%d/%m/%Y')}"
        
        for _, row in grouped.iterrows():
            report_data.append({
                'Circo': row['Circo'],
                'Período': periodo_str,
                'Faturamento Total': row['Faturamento Total'],
                'Faturamento Gestão Produtor': row['Faturamento Gestão Produtor'],
                'Taxas e Descontos': row['Taxas e Descontos'],
                'Valor Líquido': row['Valor Líquido'],
                'Total Geral': row['Faturamento Total']
            })
        
        # Armazenar dados do relatório para exportação
        self.last_report_data = report_data
        print(f"Relatório com filtro de cidades gerado com {len(report_data)} registros")
        return report_data
    
    def filter_and_generate_report_by_cities(self, selected_cidades, data_inicio, data_fim):
        """Filtrar dados e gerar relatório agrupado por cidades"""
        if not self.processed_data:
            return []
        
        # Converter para DataFrame
        df = pd.DataFrame(self.processed_data)
        
        # Filtrar por período
        try:
            df['Data Evento Parsed'] = pd.to_datetime(df['Data Evento'], format='%d/%m/%Y')
            df = df[
                (df['Data Evento Parsed'].dt.date >= data_inicio) &
                (df['Data Evento Parsed'].dt.date <= data_fim)
            ]
        except:
            pass
        
        if df.empty:
            return []
        
        # Fazer associação com cidades
        circos_cidades = circos_manager.get_all()
        
        # Adicionar coluna de cidade aos dados
        df['Cidade'] = 'Não encontrada'
        
        for index, row in df.iterrows():
            circo = row['Circo']
            data_evento_str = row['Data Evento']
            
            try:
                data_evento = datetime.strptime(data_evento_str, '%d/%m/%Y').date()
            except:
                continue
            
            for circo_cidade in circos_cidades:
                if circo_cidade['CIRCO'] == circo:
                    try:
                        data_inicio_cc = datetime.strptime(circo_cidade['DATA_INICIO'], '%d/%m/%Y').date()
                        data_fim_cc = datetime.strptime(circo_cidade['DATA_FIM'], '%d/%m/%Y').date()
                        
                        if data_inicio_cc <= data_evento <= data_fim_cc:
                            df.at[index, 'Cidade'] = circo_cidade['CIDADE']
                            break
                    except:
                        continue
        
        # Filtrar por cidades selecionadas
        df = df[df['Cidade'].isin(selected_cidades)]
        
        if df.empty:
            return []
        
        # Agrupar por cidade e somar faturamentos
        grouped = df.groupby('Cidade').agg({
            'Faturamento Total': 'sum',
            'Faturamento Gestão Produtor': 'sum',
            'Taxas e Descontos': 'sum',
            'Valor Líquido': 'sum'
        }).reset_index()
        
        # Criar dados do relatório
        report_data = []
        periodo_str = f"{data_inicio.strftime('%d/%m/%Y')} - {data_fim.strftime('%d/%m/%Y')}"
        
        for _, row in grouped.iterrows():
            report_data.append({
                'Circo': row['Cidade'],  # Usar cidade no lugar de circo para exibição
                'Período': periodo_str,
                'Faturamento Total': row['Faturamento Total'],
                'Faturamento Gestão Produtor': row['Faturamento Gestão Produtor'],
                'Taxas e Descontos': row['Taxas e Descontos'],
                'Valor Líquido': row['Valor Líquido'],
                'Total Geral': row['Faturamento Total']
            })
        
        # Armazenar dados do relatório para exportação
        self.last_report_data = report_data
        print(f"Relatório por cidades gerado com {len(report_data)} registros")
        return report_data

    def get_unique_circos(self):
        """Retorna lista de circos únicos"""
        circos = set()
        for data in self.processed_data:
            circos.add(data['Circo'])
        return sorted(list(circos))

    def create_excel_export(self, report_data):
        """Cria arquivo Excel para download"""
        output = io.BytesIO()
        
        # Converter para DataFrame para exportação
        df_export = pd.DataFrame(report_data)
        
        # Formatar valores monetários
        for col in ['Faturamento Total', 'Faturamento Gestão Produtor', 'Taxas e Descontos', 'Valor Líquido', 'Total Geral']:
            if col in df_export.columns:
                df_export[col] = df_export[col].apply(self.format_currency_display)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Relatório Sócrates Online', index=False)
            
            # Obter worksheet para formatação
            worksheet = writer.sheets['Relatório Sócrates Online']
            
            # Definir larguras específicas para cada coluna
            column_widths = {
                'A': 25,  # Circo
                'B': 25,  # Período
                'C': 20,  # Faturamento Total
                'D': 30,  # Faturamento Gestão Produtor
                'E': 20,  # Taxas e Descontos
                'F': 20,  # Valor Líquido
            }
            
            # Aplicar larguras das colunas
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Formatação adicional
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Estilo do cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar formatação ao cabeçalho
            for col in range(1, len(df_export.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Formatação das células de dados
            data_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar formatação às células de dados
            for row in range(2, len(df_export) + 2):
                for col in range(1, len(df_export.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = border
            
            # Ajustar altura das linhas
            for row in range(1, len(df_export) + 2):
                worksheet.row_dimensions[row].height = 20
        
        output.seek(0)
        return output

    def create_pdf_export(self, report_data):
        """Cria arquivo PDF para download"""
        output = io.BytesIO()
        
        # Configurar documento PDF
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Cabeçalho com marca Sócrates Online
        brand_style = ParagraphStyle(
            'BrandStyle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            alignment=1,  # Centralizado
            spaceAfter=8
        )
        
        brand_title = Paragraph("SÓCRATES ONLINE", brand_style)
        story.append(brand_title)
        
        # Título do relatório
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#764ba2'),
            alignment=1,  # Centralizado
            spaceAfter=12
        )
        
        title = Paragraph("Relatório de Faturamento por Evento", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Data de geração
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        
        date_text = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", date_style)
        story.append(date_text)
        story.append(Spacer(1, 30))
        
        if report_data:
            # Criar tabela com cabeçalhos mais curtos
            data = [['Circo', 'Período', 'Fatur. Total', 'Gestão Prod.', 'Taxas/Desc.', 'Valor Líquido']]
            
            for item in report_data:
                # Quebrar texto longo em múltiplas linhas se necessário
                circo_name = item['Circo']
                if len(circo_name) > 15:
                    circo_name = circo_name[:15] + "..."
                
                data.append([
                    circo_name,
                    item['Período'],
                    self.format_currency_display(item['Faturamento Total']),
                    self.format_currency_display(item['Faturamento Gestão Produtor']),
                    self.format_currency_display(item['Taxas e Descontos']),
                    self.format_currency_display(item['Valor Líquido'])
                ])
            
            # Calcular totais
            total_geral = sum(item['Faturamento Total'] for item in report_data)
            total_gestao = sum(item['Faturamento Gestão Produtor'] for item in report_data)
            total_taxas = sum(item['Taxas e Descontos'] for item in report_data)
            total_liquido = sum(item['Valor Líquido'] for item in report_data)
            
            data.append([
                'TOTAL',
                '',
                self.format_currency_display(total_geral),
                self.format_currency_display(total_gestao),
                self.format_currency_display(total_taxas),
                self.format_currency_display(total_liquido)
            ])
            
            # Criar tabela com larguras otimizadas para A4
            # Larguras ajustadas para caber melhor na página (6 colunas)
            table = Table(data, colWidths=[1.5*inch, 1.2*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch])
            table.setStyle(TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Dados
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.beige, colors.lightgrey]),
                
                # Linha de total
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#764ba2')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                
                # Bordas e espaçamento
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            
            story.append(table)
            
            # Adicionar resumo estatístico
            story.append(Spacer(1, 20))
            
            summary_style = ParagraphStyle(
                'SummaryStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                textColor=colors.HexColor('#333333')
            )
            
            summary_text = f"""
            <b>Resumo Executivo:</b><br/>
            • Total de Circos: {len(report_data)}<br/>
            • Faturamento Total: {self.format_currency_display(total_geral)}<br/>
            • Valor Gestão Produtor: {self.format_currency_display(total_gestao)}<br/>
            • Taxas e Descontos: {self.format_currency_display(total_taxas)}<br/>
            • Valor Líquido: {self.format_currency_display(total_liquido)}
            """
            
            summary = Paragraph(summary_text, summary_style)
            story.append(summary)
        else:
            no_data = Paragraph("Nenhum dado encontrado para os filtros selecionados.", styles['Normal'])
            story.append(no_data)
        
        # Construir PDF
        doc.build(story)
        output.seek(0)
        return output

# Instâncias globais
processor = SocratesProcessor()
circos_manager = CircosCidadesManager()

# Verificar integridade dos dados na inicialização
print("🔍 Verificando integridade dos dados de circos...")
circos_manager.verify_and_recover()

# Iniciar salvamento automático
circos_manager.start_auto_save()

print("✅ Processador inicializado com salvamento automático")

# Rotas da aplicação
@app.route('/')
def index():
    """Página principal"""
    response = app.make_response(render_template('index.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/upload', methods=['POST'])
def upload_file():
    """Upload e processamento de arquivo Excel"""
    print("=== UPLOAD REQUEST RECEBIDO ===")
    
    if 'file' not in request.files:
        print("ERRO: Nenhum arquivo no request")
        return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
    
    file = request.files['file']
    print(f"Arquivo recebido: {file.filename}")
    
    if file.filename == '':
        print("ERRO: Nome do arquivo vazio")
        return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
    
    if file and processor.allowed_file(file.filename):
        print(f"Arquivo válido, processando: {file.filename}")
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            file.save(filepath)
            print(f"Arquivo salvo em: {filepath}")
            
            # Processar arquivo
            success, message = processor.process_excel_file(filepath)
            print(f"Processamento: sucesso={success}, mensagem={message}")
            
            # Remover arquivo após processamento
            try:
                os.remove(filepath)
                print("Arquivo temporário removido")
            except:
                print("Aviso: Não foi possível remover arquivo temporário")
            
            if success:
                # Calcular estatísticas
                circos_unicos = processor.get_unique_circos()
                total_faturamento = sum([item['Faturamento Total'] for item in processor.processed_data])
                total_liquido = sum([item['Valor Líquido'] for item in processor.processed_data])
                
                # Preparar dados formatados para exibição
                display_data = []
                for item in processor.processed_data:
                    display_data.append({
                        'Circo': item['Circo'],
                        'Data Evento': item['Data Evento'],
                        'Faturamento Total': processor.format_currency_display(item['Faturamento Total']),
                        'Faturamento Gestão Produtor': processor.format_currency_display(item['Faturamento Gestão Produtor']),
                        'Taxas e Descontos': processor.format_currency_display(item['Taxas e Descontos']),
                        'Valor Líquido': processor.format_currency_display(item['Valor Líquido'])
                    })
                
                print(f"Estatísticas: {len(processor.processed_data)} registros, {len(circos_unicos)} circos")
                print(f"Display data preparado: {len(display_data)} registros")
                print(f"Primeiros 3 registros display_data: {display_data[:3] if display_data else 'Vazio'}")
                
                response_data = {
                    'success': True,
                    'message': message,
                    'stats': {
                        'total_registros': len(processor.processed_data),
                        'total_circos': len(circos_unicos),
                        'circos': circos_unicos,
                        'total_faturamento': processor.format_currency_display(total_faturamento),
                        'total_liquido': processor.format_currency_display(total_liquido)
                    },
                    'imported_data': display_data
                }
                
                print(f"Resposta final sendo enviada com imported_data: {len(response_data['imported_data'])} registros")
                return jsonify(response_data)
            else:
                print(f"ERRO no processamento: {message}")
                return jsonify({'success': False, 'message': message})
                
        except Exception as e:
            print(f"ERRO na execução: {str(e)}")
            # Tentar remover arquivo mesmo em caso de erro
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except:
                pass
            return jsonify({'success': False, 'message': f'Erro ao processar arquivo: {str(e)}'})
    else:
        print(f"ERRO: Tipo de arquivo não permitido: {file.filename}")
        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'})

@app.route('/get_circos_cidades', methods=['GET'])
def get_circos_cidades():
    """Obter dados de circos e cidades"""
    try:
        # Verificar integridade antes de retornar dados
        circos_manager.verify_and_recover()
        
        circos_data = circos_manager.get_all()
        circos_relatorio = processor.get_unique_circos() if processor.processed_data else []
        
        print(f"📊 Retornando {len(circos_data)} registros de circos-cidades")
        
        return jsonify({
            'success': True,
            'circos_cidades': circos_data,
            'circos_relatorio': circos_relatorio,
            'total_registros': len(circos_data)
        })
    except Exception as e:
        print(f"Erro ao obter dados de circos: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/add_circo_cidade', methods=['POST'])
def add_circo_cidade():
    """Adicionar novo circo/cidade"""
    try:
        data = request.get_json()
        cidade = data.get('cidade', '').strip()
        circo = data.get('circo', '').strip()
        data_inicio = data.get('data_inicio', '').strip()
        data_fim = data.get('data_fim', '').strip()
        
        if not all([cidade, circo, data_inicio, data_fim]):
            return jsonify({'success': False, 'message': 'Todos os campos são obrigatórios'})
        
        success = circos_manager.add_circo(cidade, circo, data_inicio, data_fim)
        
        if success:
            return jsonify({'success': True, 'message': 'Circo adicionado com sucesso'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao adicionar circo'})
            
    except Exception as e:
        print(f"Erro ao adicionar circo: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/update_circo_cidade', methods=['POST'])
def update_circo_cidade():
    """Atualizar circo/cidade existente"""
    try:
        data = request.get_json()
        index = data.get('index')
        cidade = data.get('cidade', '').strip()
        circo = data.get('circo', '').strip()
        data_inicio = data.get('data_inicio', '').strip()
        data_fim = data.get('data_fim', '').strip()
        
        if index is None or not all([cidade, circo, data_inicio, data_fim]):
            return jsonify({'success': False, 'message': 'Todos os campos são obrigatórios'})
        
        success = circos_manager.update_circo(index, cidade, circo, data_inicio, data_fim)
        
        if success:
            return jsonify({'success': True, 'message': 'Circo atualizado com sucesso'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao atualizar circo'})
            
    except Exception as e:
        print(f"Erro ao atualizar circo: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/delete_circo_cidade', methods=['POST'])
def delete_circo_cidade():
    """Deletar circo/cidade"""
    try:
        data = request.get_json()
        index = data.get('index')
        
        if index is None:
            return jsonify({'success': False, 'message': 'Índice não fornecido'})
        
        success = circos_manager.delete_circo(index)
        
        if success:
            return jsonify({'success': True, 'message': 'Circo removido com sucesso'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao remover circo'})
            
    except Exception as e:
        print(f"Erro ao deletar circo: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/confirm_circos', methods=['POST'])
def confirm_circos():
    """Confirmar cadastro de circos e liberar próximas etapas"""
    try:
        return jsonify({'success': True, 'message': 'Cadastro de circos confirmado'})
    except Exception as e:
        print(f"Erro ao confirmar circos: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/status_dados', methods=['GET'])
def status_dados():
    """Verificar status dos dados para debug"""
    try:
        csv_file = circos_manager.csv_file
        csv_path = os.path.abspath(csv_file)
        
        status = {
            'arquivo_existe': os.path.exists(csv_file),
            'caminho_absoluto': csv_path,
            'tamanho_arquivo': os.path.getsize(csv_file) if os.path.exists(csv_file) else 0,
            'dados_em_memoria': len(circos_manager.data),
            'backup_existe': os.path.exists(os.path.join(circos_manager.backup_folder, f"{os.path.basename(csv_file)}.backup")),
            'pasta_backup': circos_manager.backup_folder,
            'backups_na_pasta': len([f for f in os.listdir(circos_manager.backup_folder) if f.startswith(os.path.basename(csv_file)) and 'backup' in f]) if os.path.exists(circos_manager.backup_folder) else 0,
            'registros_atuais': circos_manager.get_all()
        }
        
        # Tentar verificar integridade
        integridade_ok = circos_manager.verify_and_recover()
        status['integridade_ok'] = integridade_ok
        
        return jsonify({
            'success': True,
            'status': status
        })
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Erro ao verificar status: {str(e)}'
        })

@app.route('/associate_cities_to_data', methods=['GET'])
def associate_cities_to_data():
    """Associar cidades aos dados importados baseado em circo e período"""
    try:
        if not processor.processed_data:
            return jsonify({'success': False, 'message': 'Nenhum dado importado encontrado'})
        
        # Obter dados de circos e cidades
        circos_cidades = circos_manager.get_all()
        
        if not circos_cidades:
            return jsonify({'success': False, 'message': 'Nenhum cadastro de circo-cidade encontrado'})
        
        # Fazer associação
        associated_data = []
        
        for item in processor.processed_data:
            circo = item['Circo']
            data_evento_str = item['Data Evento']
            
            # Converter data do evento para comparação
            try:
                data_evento = datetime.strptime(data_evento_str, '%d/%m/%Y').date()
            except:
                continue
            
            # Procurar cidade correspondente
            cidade_encontrada = None
            
            for circo_cidade in circos_cidades:
                if circo_cidade['CIRCO'] == circo:
                    try:
                        # Converter datas do CSV
                        data_inicio = datetime.strptime(circo_cidade['DATA_INICIO'], '%d/%m/%Y').date()
                        data_fim = datetime.strptime(circo_cidade['DATA_FIM'], '%d/%m/%Y').date()
                        
                        # Debug: verificar consistência das datas
                        if data_inicio > data_fim:
                            print(f"⚠️ AVISO: Datas inconsistentes para {circo} em {circo_cidade['CIDADE']}: {data_inicio} > {data_fim}")
                            continue
                        
                        # Verificar se a data do evento está no período
                        if data_inicio <= data_evento <= data_fim:
                            cidade_encontrada = circo_cidade['CIDADE']
                            print(f"✅ Associação encontrada: {circo} → {cidade_encontrada} (evento: {data_evento})")
                            break
                    except Exception as e:
                        print(f"⚠️ Erro ao processar datas para {circo}: {e}")
                        continue
            
            # Adicionar registro com cidade
            associated_item = {
                'Circo': circo,
                'Cidade': cidade_encontrada if cidade_encontrada else 'Não encontrada',
                'Data Evento': data_evento_str,
                'Faturamento Total': processor.format_currency_display(item['Faturamento Total']),
                'Faturamento Gestão Produtor': processor.format_currency_display(item['Faturamento Gestão Produtor']),
                'Taxas e Descontos': processor.format_currency_display(item['Taxas e Descontos']),
                'Valor Líquido': processor.format_currency_display(item['Valor Líquido'])
            }
            
            associated_data.append(associated_item)
        
        print(f"Associação realizada: {len(associated_data)} registros processados")
        
        return jsonify({
            'success': True,
            'associated_data': associated_data,
            'total_records': len(associated_data)
        })
        
    except Exception as e:
        print(f"Erro na associação: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/generate_report', methods=['POST'])
def generate_report():
    """Gerar relatório com filtros"""
    try:
        data = request.get_json()
        
        tipo_filtro = data.get('tipo_filtro', 'circo')
        data_inicio = datetime.strptime(data.get('data_inicio'), '%Y-%m-%d').date()
        data_fim = datetime.strptime(data.get('data_fim'), '%Y-%m-%d').date()
        
        if data_inicio > data_fim:
            return jsonify({'success': False, 'message': 'Data inicial deve ser anterior à data final'})
        
        if tipo_filtro == 'circo':
            # Relatório por circo (modo tradicional)
            selected_circos = data.get('circos', [])
            if not selected_circos:
                return jsonify({'success': False, 'message': 'Selecione pelo menos um circo'})
            
            report_data = processor.filter_and_generate_report(selected_circos, data_inicio, data_fim)
        else:
            # Relatório por cidade
            selected_cidades = data.get('cidades', [])
            if not selected_cidades:
                return jsonify({'success': False, 'message': 'Selecione pelo menos uma cidade'})
            
            report_data = processor.filter_and_generate_report_by_cities(selected_cidades, data_inicio, data_fim)
        
        if not report_data:
            return jsonify({'success': False, 'message': 'Nenhum dado encontrado para os filtros selecionados'})
        
        # Garantir que os dados são salvos para exportação
        processor.last_report_data = report_data
        
        # Calcular totais
        total_geral = sum([item['Faturamento Total'] for item in report_data])
        total_gestao = sum([item['Faturamento Gestão Produtor'] for item in report_data])
        total_taxas = sum([item['Taxas e Descontos'] for item in report_data])
        total_liquido = sum([item['Valor Líquido'] for item in report_data])
        
        # Formatar dados para exibição
        display_data = []
        for item in report_data:
            display_data.append({
                'Circo': item['Circo'],  # Este campo contém circo ou cidade dependendo do tipo_filtro
                'Período': item['Período'],
                'Faturamento Total': processor.format_currency_display(item['Faturamento Total']),
                'Faturamento Gestão Produtor': processor.format_currency_display(item['Faturamento Gestão Produtor']),
                'Taxas e Descontos': processor.format_currency_display(item['Taxas e Descontos']),
                'Valor Líquido': processor.format_currency_display(item['Valor Líquido'])
            })
        
        # Criar gráficos (apenas se Plotly estiver disponível)
        fig_pie = fig_comp = None
        
        if PLOTLY_AVAILABLE:
            try:
                
                # Gráfico de pizza Gestão vs Líquido
                gestao_liquido_data = {
                    'Tipo': ['Valor líquido em dinheiro', 'Valor Líquido'],
                    'Valor': [total_gestao, total_liquido]
                }
                
                fig_pie = px.pie(
                    gestao_liquido_data,
                    values='Valor',
                    names='Tipo',
                    title='Distribuição: Valor líquido em dinheiro vs Valor Líquido',
                    color_discrete_sequence=['#ff7f0e', '#28a745']
                )
                
                # Gráfico comparativo
                comparison_data = []
                label_type = 'Cidade' if tipo_filtro == 'cidade' else 'Circo'
                
                for item in report_data:
                    comparison_data.extend([
                        {label_type: item['Circo'], 'Tipo': 'Valor líquido em dinheiro', 'Valor': item['Faturamento Gestão Produtor']},
                        {label_type: item['Circo'], 'Tipo': 'Valor Líquido', 'Valor': item['Valor Líquido']}
                    ])
                
                chart_title = f'Comparativo Valor líquido em dinheiro vs Valor Líquido por {label_type}'
                
                fig_comp = px.bar(
                    comparison_data,
                    x=label_type,
                    y='Valor',
                    color='Tipo',
                    title=chart_title,
                    barmode='group'
                )
                
                print("✅ Gráficos gerados com sucesso")
                
            except Exception as e:
                print(f"⚠️ Erro ao gerar gráficos: {e}")
                fig_pie = fig_comp = None
        
        # Preparar resposta
        response_data = {
            'success': True,
            'data': display_data,
            'tipo_filtro': tipo_filtro,  # Incluir tipo de filtro na resposta
            'stats': {
                'total_geral': processor.format_currency_display(total_geral),
                'total_gestao': processor.format_currency_display(total_gestao),
                'total_taxas': processor.format_currency_display(total_taxas),
                'total_liquido': processor.format_currency_display(total_liquido),
                'total_circos': len(report_data)
            }
        }
        
        # Adicionar gráficos apenas se Plotly estiver disponível e gráficos foram criados
        if PLOTLY_AVAILABLE and fig_pie and fig_comp:
            try:
                response_data['charts'] = {
                    'pie': json.dumps(fig_pie, cls=plotly.utils.PlotlyJSONEncoder),
                    'comparison': json.dumps(fig_comp, cls=plotly.utils.PlotlyJSONEncoder)
                }
                print("✅ Gráficos serializados com sucesso")
            except Exception as e:
                print(f"⚠️ Erro ao serializar gráficos: {e}")
                response_data['charts'] = {'pie': '{}', 'comparison': '{}'}
        else:
            print("⚠️ Gráficos não disponíveis - usando dados vazios")
            response_data['charts'] = {'pie': '{}', 'comparison': '{}'}
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao gerar relatório: {str(e)}'})

@app.route('/export/<export_type>')
def export_report(export_type):
    """Exportar relatório"""
    try:
        # Usar dados do último relatório gerado
        if not hasattr(processor, 'last_report_data') or not processor.last_report_data:
            flash('Gere um relatório primeiro antes de exportar', 'warning')
            return redirect(url_for('index'))
        
        if export_type == 'excel':
            excel_data = processor.create_excel_export(processor.last_report_data)
            filename = f"relatorio_socrates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                excel_data,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        elif export_type == 'pdf':
            pdf_data = processor.create_pdf_export(processor.last_report_data)
            filename = f"relatorio_socrates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            return send_file(
                pdf_data,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        
        else:
            flash('Tipo de exportação inválido', 'error')
            return redirect(url_for('index'))
            
    except Exception as e:
        flash(f'Erro ao exportar: {str(e)}', 'error')
        return redirect(url_for('index'))

# Filtro para templates
@app.template_filter('currency')
def currency_filter(value):
    """Filtro para formatar moeda nos templates"""
    return processor.format_currency_display(value)

def shutdown_handler():
    """Limpar recursos antes de encerrar a aplicação"""
    try:
        print("\n🛑 Encerrando aplicação...")
        
        # Parar salvamento automático
        circos_manager.stop_auto_save()
        
        # Fazer backup final
        if circos_manager.data:
            print("💾 Fazendo salvamento final...")
            circos_manager.save_data()
        
        print("✅ Encerramento limpo concluído")
    except Exception as e:
        print(f"⚠️ Erro no encerramento: {e}")

if __name__ == '__main__':
    try:
        # Configurar para desenvolvimento
        print("🎪 Sócrates Online iniciando...")
        print("🌐 Acesse: http://localhost:5000")
        print("🛑 Para parar: Ctrl+C")
        
        # Configurar handler de sinal para encerramento limpo
        import signal
        def signal_handler(sig, frame):
            shutdown_handler()
            import sys
            sys.exit(0)
        
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        
        app.run(debug=False, host='0.0.0.0', port=5000)
        
    except KeyboardInterrupt:
        shutdown_handler()
    except Exception as e:
        print(f"❌ Erro fatal: {e}")
        shutdown_handler()
